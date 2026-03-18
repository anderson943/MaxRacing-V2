"""
MaxRacing AI Knowledge Base Builder (v2) — Catalog + Instructions (ALL INFO)

What this v2 does (in one run):
1) Unzips BOTH:
   - CATALOGO SUPORTES (XLSX workbooks with per-brand sheets + embedded images + BOM tables)
   - INSTRUÇÃO DE MONTAGEM (PDF / DOCX / images)
2) Indexes EVERYTHING useful:
   - Catalog: bike model(s), damper model (MAX10/MAX20 when CAPA is present), variant, BOM, embedded images,
             and ALSO extracts "notes" text blocks from each sheet.
   - Instructions: PDF/DOCX/images (flat or nested), infers model name from filename and (optionally) brand.
3) Removes internal codes like "10 - 001" from customer-facing fields.
4) Matches catalog entries → instruction files even when instruction ZIP is FLAT (no brand folders).
   - Matching is token-based + fuzzy overlap.
   - Prefers English instruction files if filename contains "INGLES"/"ENGLISH".
5) Outputs:
   - JSONL knowledge base ready for embeddings/RAG
   - assets/ folder containing extracted images + copied PDFs/DOCX/images
   - match_report.json for debugging unmatched items

Run:
  python maxracing_kb_v2.py

"""

import os
import re
import json
import zipfile
import shutil
import hashlib
import unicodedata
from typing import Any, Dict, List, Tuple, Optional

import openpyxl


# -----------------------------
# CONFIG
# -----------------------------
ZIP_CATALOGO = "/mnt/data/CATALOGO SUPORTES-20260304T174307Z-3-001.zip"
ZIP_INSTRUCOES = "/mnt/data/INSTRUÇÃO DE MONTAGEM_1-20260304T174424Z-3-001.zip"

OUT_DIR = "/mnt/data/maxracing_ai_kb_v2"
ASSETS_DIR = os.path.join(OUT_DIR, "assets")
JSONL_PATH = os.path.join(OUT_DIR, "maxracing_kb_v2.jsonl")
MATCH_REPORT = os.path.join(OUT_DIR, "match_report_v2.json")

# Internal code patterns to remove (customer-facing)
# Examples: "10 - 001", "20-002", "AJP - 10 - 001"
RE_INTERNAL_CODE = re.compile(r"\b([A-Z]{2,}\s*-\s*)?\d{2}\s*-\s*\d{3}\b", re.IGNORECASE)

# Ignore template/non-data sheets
IGNORE_SHEETS_EXACT = {"CAPA", "MODELO"}
IGNORE_SHEETS_REGEX = [
    re.compile(r"^PLANILHA\d*$", re.IGNORECASE),
    re.compile(r"^SHEET\d*$", re.IGNORECASE),
]

# BOM headers (Portuguese)
BOM_HEADERS = ["ITEM", "COR", "PARAFUSO", "QTDE"]  # may appear anywhere on sheet

# Instruction file extensions to index (v2 includes DOCX)
INSTR_EXTS = (".pdf", ".docx", ".png", ".jpg", ".jpeg", ".webp")

# Match thresholds
MATCH_THRESHOLD_ATTACH = 0.48     # attach instructions to catalog entry
MATCH_THRESHOLD_STRONG = 0.62     # strong match

# Prefer English instruction files if present
ENGLISH_HINTS = ("INGLES", "ENGLISH", "_EN", "-EN", " EN ")


# -----------------------------
# BASIC HELPERS
# -----------------------------
def ensure_clean_dir(path: str):
    if os.path.exists(path):
        shutil.rmtree(path)
    os.makedirs(path, exist_ok=True)

def safe_mkdir(path: str):
    os.makedirs(path, exist_ok=True)

def sha1_id(*parts: str) -> str:
    raw = "||".join(parts).encode("utf-8", errors="ignore")
    return hashlib.sha1(raw).hexdigest()

def clean_text(s: Any) -> str:
    """Customer-facing cleaning: remove internal codes, trim excess punctuation/spaces."""
    if s is None:
        return ""
    txt = str(s).strip()
    txt = RE_INTERNAL_CODE.sub("", txt)
    txt = re.sub(r"\s{2,}", " ", txt).strip(" -–—|")
    return txt.strip()

def normalize_key(s: str) -> str:
    """Matching key: remove accents/punct, uppercase, collapse spaces."""
    s = clean_text(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s

def split_models(model_line: str) -> List[str]:
    """Split 'A / B / C' into models list."""
    if not model_line:
        return []
    parts = [p.strip() for p in str(model_line).split("/") if p.strip()]
    return parts if parts else [str(model_line).strip()]

def unzip(zip_path: str, dest_dir: str) -> str:
    safe_mkdir(dest_dir)
    with zipfile.ZipFile(zip_path, "r") as z:
        z.extractall(dest_dir)
    return dest_dir

def find_files(root: str, exts: Tuple[str, ...]) -> List[str]:
    out = []
    for r, _, files in os.walk(root):
        for f in files:
            if f.lower().endswith(exts):
                out.append(os.path.join(r, f))
    return sorted(out)

def is_ignored_sheet(name: str) -> bool:
    if name in IGNORE_SHEETS_EXACT:
        return True
    for rx in IGNORE_SHEETS_REGEX:
        if rx.match(name.strip()):
            return True
    return False

def copy_asset(src_path: str, kind: str, brand: str = "UNKNOWN") -> str:
    """Copy any file to assets folder and return relative path."""
    base = os.path.basename(src_path)
    base = re.sub(r"[^a-zA-Z0-9_\-\.]+", "_", base)
    out_name = f"{kind}__{brand}__{base}"
    out_path = os.path.join(ASSETS_DIR, out_name)
    shutil.copy2(src_path, out_path)
    return f"assets/{out_name}"


# -----------------------------
# CATALOG: CAPA MAPPING
# -----------------------------
def read_capa_mapping(wb) -> Dict[str, Dict[str, str]]:
    """
    If CAPA exists, commonly:
      B = MAX10/MAX20
      D = "PR 4" / "PR 5" etc or sheet key
    We'll map D -> { damper_model, variant_clean }.
    """
    if "CAPA" not in wb.sheetnames:
        return {}
    ws = wb["CAPA"]
    mapping = {}
    for r in range(1, 500):
        b = ws[f"B{r}"].value
        d = ws[f"D{r}"].value
        if not b or not d:
            continue
        damper_model = clean_text(b).upper()  # MAX10/MAX20
        key = str(d).strip()
        mapping[key] = {
            "damper_model": damper_model,
            "variant_raw": key,
            "variant_clean": clean_text(key),
        }
    return mapping


# -----------------------------
# CATALOG: BOM EXTRACT (ROBUST)
# -----------------------------
def find_bom_table(ws) -> Tuple[List[Dict[str, Any]], Optional[Dict[str, int]]]:
    """
    v2 BOM finder:
    - Searches the whole sheet for a row containing all headers (ITEM/COR/PARAFUSO/QTDE) in any columns.
    - Returns BOM rows and the column index mapping.
    """
    header_row = None
    col_map = None

    # Scan rows for headers anywhere
    for row_idx in range(1, ws.max_row + 1):
        row_vals = []
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            row_vals.append(str(v).strip().upper() if v is not None else "")

        # Find positions of each header in the row
        positions = {}
        for hdr in BOM_HEADERS:
            try:
                pos = row_vals.index(hdr)
                positions[hdr] = pos + 1  # convert to 1-based col index
            except ValueError:
                positions[hdr] = None

        if all(positions[h] is not None for h in BOM_HEADERS):
            header_row = row_idx
            col_map = {h: positions[h] for h in BOM_HEADERS}
            break

    if header_row is None or col_map is None:
        return [], None

    bom_rows = []
    # Read until ITEM column is blank (or hit too many blank rows)
    blank_streak = 0
    for r in range(header_row + 1, ws.max_row + 1):
        item = ws.cell(row=r, column=col_map["ITEM"]).value
        cor = ws.cell(row=r, column=col_map["COR"]).value
        parafuso = ws.cell(row=r, column=col_map["PARAFUSO"]).value
        qtde = ws.cell(row=r, column=col_map["QTDE"]).value

        if item is None or str(item).strip() == "":
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0

        bom_rows.append({
            "item": clean_text(item),
            "color": clean_text(cor),
            "fastener": clean_text(parafuso),
            "qty": qtde if qtde is not None else None
        })

    return bom_rows, col_map


# -----------------------------
# CATALOG: NOTES / TEXT BLOCKS
# -----------------------------
def extract_text_notes(ws) -> List[str]:
    """
    v2 notes extractor:
    - Captures meaningful text from cells outside BOM rows.
    - Filters out single-character noise and obvious headers.
    - Great for sheet-specific mounting notes.
    """
    notes = []
    # Gather all text-ish cells
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, (int, float)):
                continue
            txt = clean_text(v)
            if not txt:
                continue
            # Filter obvious table headers / repeated junk
            up = txt.upper()
            if up in set(BOM_HEADERS) or up in {"ITEM", "COR", "PARAFUSO", "QTDE"}:
                continue
            if len(txt) <= 2:
                continue
            notes.append(txt)

    # De-dup while preserving order
    seen = set()
    out = []
    for n in notes:
        key = normalize_key(n)
        if key in seen:
            continue
        seen.add(key)
        out.append(n)

    # Keep notes reasonable (avoid dumping entire sheets of repeated labels)
    return out[:120]


# -----------------------------
# CATALOG: IMAGE EXTRACTION
# -----------------------------
def extract_images_from_sheet(wb, sheet_name: str, brand: str) -> List[str]:
    ws = wb[sheet_name]
    images = getattr(ws, "_images", []) or []
    saved = []
    for idx, img in enumerate(images, start=1):
        try:
            data = img._data()
        except Exception:
            continue
        fname = f"catalog__{brand}__{sheet_name}__img{idx}.png"
        fname = re.sub(r"[^a-zA-Z0-9_\-\.]+", "_", fname)
        out_path = os.path.join(ASSETS_DIR, fname)
        with open(out_path, "wb") as f:
            f.write(data)
        saved.append(f"assets/{fname}")
    return saved


# -----------------------------
# INDEX: CATALOG XLSX
# -----------------------------
def index_catalog_xlsx(xlsx_path: str) -> List[Dict[str, Any]]:
    brand = os.path.splitext(os.path.basename(xlsx_path))[0].strip()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    capa_map = read_capa_mapping(wb)
    records = []

    for sheet_name in wb.sheetnames:
        if is_ignored_sheet(sheet_name):
            continue

        ws = wb[sheet_name]

        # Bike model line in A1 usually; fallback to sheet name
        a1 = ws["A1"].value
        model_line = clean_text(a1) if a1 else clean_text(sheet_name)
        motorcycle_models = split_models(model_line) if model_line else [clean_text(sheet_name)]
        motorcycle_models = [m for m in motorcycle_models if m] or [clean_text(sheet_name)]

        # CAPA mapping by sheet_name (common); if not found, leave null
        cap = capa_map.get(sheet_name)
        damper_model = cap["damper_model"] if cap else None
        variant = cap["variant_clean"] if cap else clean_text(sheet_name)

        # BOM (robust)
        bom, bom_col_map = find_bom_table(ws)

        # Notes (sheet text blocks)
        notes = extract_text_notes(ws)

        # Embedded images
        images = extract_images_from_sheet(wb, sheet_name, brand)

        rec_id = sha1_id("catalog", brand, sheet_name, os.path.basename(xlsx_path))
        model_keys = [normalize_key(m) for m in motorcycle_models]

        records.append({
            "id": rec_id,
            "type": "support_catalog_entry",
            "brand": brand,
            "source_workbook": os.path.basename(xlsx_path),
            "sheet": sheet_name,

            "damper_model": damper_model,     # MAX10/MAX20 if CAPA present
            "variant": variant,               # PR 4 / PR 5 / etc or sheet name

            "motorcycle_models": motorcycle_models,
            "motorcycle_model_keys": model_keys,   # internal matching only

            "bill_of_materials": bom,
            "sheet_notes": notes,

            "images": images,
            "instructions": [],

            "source": {"file": os.path.basename(xlsx_path), "sheet": sheet_name}
        })

    return records


# -----------------------------
# INSTRUCTIONS: BRAND/MODEL INFERENCE
# -----------------------------
def infer_brand_from_filename(filename: str, known_brands: List[str]) -> str:
    """
    Instructions ZIP may be flat. We'll try to detect brand by checking if filename contains a known brand token.
    Otherwise UNKNOWN.
    """
    base = normalize_key(os.path.splitext(os.path.basename(filename))[0])
    for b in known_brands:
        nb = normalize_key(b)
        if nb and nb in base.split():
            return b
        # also allow substring match for brands like "HONDA", "YAMAHA"
        if nb and nb in base:
            return b
    return "UNKNOWN"

def infer_model_from_filename(path: str) -> str:
    """
    Extract model text from filename. Remove brand-ish tokens later during matching if needed.
    """
    base = os.path.splitext(os.path.basename(path))[0]
    base = base.replace("_", " ").replace("-", " ")
    base = re.sub(r"\s{2,}", " ", base).strip()
    base = clean_text(base)
    return base

def is_english_instruction(path: str) -> bool:
    up = normalize_key(os.path.basename(path))
    for h in ENGLISH_HINTS:
        if normalize_key(h) in up:
            return True
    return False


# -----------------------------
# INDEX: INSTRUCTIONS (PDF/DOCX/IMG)
# -----------------------------
def index_instructions(instructions_root: str, known_brands: List[str]) -> List[Dict[str, Any]]:
    files = find_files(instructions_root, INSTR_EXTS)
    recs = []

    for fp in files:
        model = infer_model_from_filename(fp)
        brand = infer_brand_from_filename(fp, known_brands)

        asset_rel = copy_asset(fp, kind="instr", brand=brand)
        rec_id = sha1_id("instruction", brand, model, os.path.basename(fp))

        recs.append({
            "id": rec_id,
            "type": "assembly_instruction",
            "brand": brand,
            "motorcycle_model": model,
            "motorcycle_model_key": normalize_key(model),
            "file": asset_rel,
            "file_type": os.path.splitext(fp)[1].lower().lstrip("."),
            "language_hint": "EN" if is_english_instruction(fp) else "UNKNOWN",
            "source": {"original_path": fp}
        })

    return recs


# -----------------------------
# MATCHING: CATALOG ↔ INSTRUCTIONS (MODEL-FIRST)
# -----------------------------
def token_overlap_score(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    ta = set(a.split())
    tb = set(b.split())
    if not ta or not tb:
        return 0.0
    overlap = len(ta & tb)
    return overlap / max(len(ta), len(tb))

def best_instruction_matches_for_catalog(
    catalog_entry: Dict[str, Any],
    instructions: List[Dict[str, Any]],
    limit: int = 5
) -> List[Dict[str, Any]]:
    """
    Score instructions by:
    - model token overlap against each catalog model key
    - slight bonus if brand matches (when instruction brand is not UNKNOWN)
    - bonus if English (prefer in ties)
    """
    best: List[Tuple[float, Dict[str, Any]]] = []

    model_keys = catalog_entry.get("motorcycle_model_keys", []) or []
    cat_brand = normalize_key(catalog_entry.get("brand", ""))

    for ins in instructions:
        ins_key = ins["motorcycle_model_key"]
        score = 0.0
        for mk in model_keys:
            score = max(score, token_overlap_score(mk, ins_key))

        if score <= 0:
            continue

        # brand bonus (only if instruction has a real brand)
        ins_brand = normalize_key(ins.get("brand", "UNKNOWN"))
        if ins_brand and ins_brand != "UNKNOWN" and ins_brand == cat_brand:
            score += 0.06

        # English tie-break bonus
        if ins.get("language_hint") == "EN":
            score += 0.02

        best.append((score, ins))

    best.sort(key=lambda x: x[0], reverse=True)
    out = []
    for score, ins in best[:limit]:
        out.append({
            "instruction_id": ins["id"],
            "file": ins["file"],
            "file_type": ins["file_type"],
            "motorcycle_model": ins["motorcycle_model"],
            "brand": ins.get("brand", "UNKNOWN"),
            "language_hint": ins.get("language_hint", "UNKNOWN"),
            "match_score": round(score, 3),
        })
    return out

def attach_instructions_to_catalog(
    catalog_records: List[Dict[str, Any]],
    instruction_records: List[Dict[str, Any]]
) -> Dict[str, Any]:
    report = {"matched": [], "unmatched_catalog": [], "unmatched_instructions": []}
    used_instr_ids = set()

    for c in catalog_records:
        candidates = best_instruction_matches_for_catalog(c, instruction_records, limit=6)

        # Attach those above threshold; prefer English if multiple similar
        attached = []
        for cand in candidates:
            if cand["match_score"] < MATCH_THRESHOLD_ATTACH:
                continue
            attached.append(cand)
            used_instr_ids.add(cand["instruction_id"])

        # If multiple: sort by score then EN preference
        attached.sort(key=lambda x: (x["match_score"], 1 if x["language_hint"] == "EN" else 0), reverse=True)

        # Keep top 3 (most relevant)
        c["instructions"] = attached[:3]

        if c["instructions"]:
            report["matched"].append({
                "catalog_id": c["id"],
                "brand": c["brand"],
                "catalog_models": c["motorcycle_models"],
                "attached": c["instructions"]
            })
        else:
            report["unmatched_catalog"].append({
                "catalog_id": c["id"],
                "brand": c["brand"],
                "catalog_models": c["motorcycle_models"]
            })

    for ins in instruction_records:
        if ins["id"] not in used_instr_ids:
            report["unmatched_instructions"].append({
                "instruction_id": ins["id"],
                "brand": ins["brand"],
                "model": ins["motorcycle_model"],
                "file": ins["file"],
                "language_hint": ins.get("language_hint", "UNKNOWN")
            })

    return report


# -----------------------------
# RUN
# -----------------------------
def run():
    ensure_clean_dir(OUT_DIR)
    safe_mkdir(ASSETS_DIR)

    # unzip both zips
    cat_root = unzip(ZIP_CATALOGO, os.path.join(OUT_DIR, "unzipped_catalog"))
    ins_root = unzip(ZIP_INSTRUCOES, os.path.join(OUT_DIR, "unzipped_instructions"))

    # index ALL xlsx in catalog (v2 does NOT exclude any workbook)
    xlsx_files = find_files(cat_root, (".xlsx",))
    if not xlsx_files:
        raise FileNotFoundError("No .xlsx files found inside CATALOGO SUPORTES zip.")

    catalog_records: List[Dict[str, Any]] = []
    for xp in xlsx_files:
        # Skip temporary Excel lock files
        if os.path.basename(xp).startswith("~$"):
            continue
        catalog_records.extend(index_catalog_xlsx(xp))

    # known brands from catalog workbook names
    known_brands = sorted(list({c["brand"] for c in catalog_records if c.get("brand")}))

    # index instructions (pdf/docx/images), flat or nested
    instruction_records = index_instructions(ins_root, known_brands=known_brands)

    # match instructions to catalog
    match_report = attach_instructions_to_catalog(catalog_records, instruction_records)

    # Write JSONL:
    # - catalog entries (with instructions attached)
    # - instruction entries standalone (for RAG/embeddings)
    with open(JSONL_PATH, "w", encoding="utf-8") as f:
        for rec in catalog_records:
            rec_out = dict(rec)
            # Remove internal matching keys from customer-facing output
            rec_out.pop("motorcycle_model_keys", None)
            f.write(json.dumps(rec_out, ensure_ascii=False) + "\n")
        for rec in instruction_records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    with open(MATCH_REPORT, "w", encoding="utf-8") as f:
        json.dump(match_report, f, ensure_ascii=False, indent=2)

    print("✅ MaxRacing KB v2 built successfully")
    print(f"Catalog records: {len(catalog_records)}")
    print(f"Instruction records: {len(instruction_records)}")
    print(f"JSONL: {JSONL_PATH}")
    print(f"Assets: {ASSETS_DIR}")
    print(f"Match report: {MATCH_REPORT}")


if __name__ == "__main__":
    run()
